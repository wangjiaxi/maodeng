import http.server
import json
import os
import openpyxl

DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(DIR, 'group.xlsx')


class Handler(http.server.BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            self.send_file(os.path.join(DIR, 'index.html'), 'text/html; charset=utf-8')
        elif self.path == '/api/data':
            self.handle_read()
        else:
            self.send_error(404)

    def do_POST(self):
        if self.path == '/api/data':
            self.handle_write()
        else:
            self.send_error(404)

    def send_file(self, path, content_type):
        with open(path, 'rb') as f:
            data = f.read()
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_read(self):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            ids = []
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                val = row[0]
                if val is not None and str(val).strip():
                    ids.append(str(val).strip())
            self.send_json({'ok': True, 'ids': ids})
        except Exception as e:
            self.send_json({'ok': False, 'error': str(e)})

    def handle_write(self):
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length))
        ids = body.get('ids', [])

        if not ids:
            self.send_json({'ok': False, 'error': 'no ids'})
            return

        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active

            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=1, value=None)

            for i, uid in enumerate(ids):
                ws.cell(row=i + 2, column=1, value=uid)

            wb.save(EXCEL_PATH)
            self.send_json({'ok': True, 'count': len(ids)})
        except Exception as e:
            self.send_json({'ok': False, 'error': str(e)})

    def send_json(self, data):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        print(f'[server] {args[0]}')


if __name__ == '__main__':
    port = 8765
    with http.server.HTTPServer(('127.0.0.1', port), Handler) as server:
        print(f'服务已启动: http://localhost:{port}')
        print(f'Excel 文件: {EXCEL_PATH}')
        server.serve_forever()
